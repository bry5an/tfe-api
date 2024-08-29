import pandas as pd
import networkx as nx
from openpyxl import load_workbook
from openpyxl.styles import Font

# Step 1: Read the Excel file
df = pd.read_excel('workspaces.xlsx')

# Step 2: Create a directed graph
G = nx.DiGraph()

# Add nodes
for index, row in df.iterrows():
    G.add_node(row['WorkspaceName'])

# Add edges based on remote workspaces
for index, row in df.iterrows():
    if pd.notna(row['RemoteWorkspace']):
        remote_workspaces = row['RemoteWorkspace'].split(',')
        for remote_workspace in remote_workspaces:
            G.add_edge(remote_workspace.strip(), row['WorkspaceName'])

# Step 3: Detect cycles and assign unique values
cycles = list(nx.simple_cycles(G))
cycle_workspaces = set()
cycle_order_dict = {}

for i, cycle in enumerate(cycles):
    for workspace in cycle:
        cycle_workspaces.add(workspace)
        cycle_order_dict[workspace] = f'cycle-{i+1}'

# Remove cycle workspaces from the graph
G.remove_nodes_from(cycle_workspaces)

# Step 4: Calculate the depth of each workspace
depth_dict = {}
for node in nx.topological_sort(G):
    predecessors = list(G.predecessors(node))
    if not predecessors:
        depth_dict[node] = 0
    else:
        depth_dict[node] = max(depth_dict[pred] for pred in predecessors) + 1

# Assign group numbers based on depth
group_dict = {workspace: depth + 1 for workspace, depth in depth_dict.items()}

# Combine the two dictionaries
group_dict.update(cycle_order_dict)

# Step 5: Add the group number to the DataFrame
df['MigrationGroup'] = df['WorkspaceName'].map(group_dict)

# Save the updated DataFrame back to the Excel file
df.to_excel('workspaces_with_group.xlsx', index=False)

# Step 6: Highlight remote workspace names by changing text color
wb = load_workbook('workspaces_with_group.xlsx')
ws = wb.active

# List of unique colors
colors = [
    "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF", "800000", "008000", "000080",
    "808000", "800080", "008080", "C0C0C0", "808080", "FFA500", "A52A2A", "8A2BE2", "5F9EA0",
    "D2691E", "FF7F50", "6495ED", "DC143C", "00FFFF", "00008B", "008B8B", "B8860B", "A9A9A9",
    "006400", "BDB76B", "8B008B", "556B2F", "FF8C00", "9932CC", "8B0000", "E9967A", "8FBC8F",
    "483D8B", "2F4F4F", "00CED1", "9400D3", "FF1493", "00BFFF", "696969", "1E90FF", "B22222",
    "FFFAF0", "228B22", "FF00FF", "DCDCDC", "F8F8FF", "FFD700", "DAA520", "808080", "008000",
    "ADFF2F", "F0FFF0", "FF69B4", "CD5C5C", "4B0082", "FFFFF0", "F0E68C", "E6E6FA", "FFF0F5",
    "7CFC00", "FFFACD", "ADD8E6", "F08080", "E0FFFF", "FAFAD2", "D3D3D3", "90EE90", "FFB6C1",
    "FFA07A", "20B2AA", "87CEFA", "778899", "B0C4DE", "FFFFE0", "00FF00", "32CD32", "FAF0E6",
    "FF00FF", "800000", "66CDAA", "0000CD", "BA55D3", "9370DB", "3CB371", "7B68EE", "00FA9A",
    "48D1CC", "C71585", "191970", "F5FFFA", "FFE4E1", "FFE4B5", "FFDEAD", "000080", "FDF5E6",
    "808000", "6B8E23", "FFA500", "FF4500", "DA70D6", "EEE8AA", "98FB98", "AFEEEE", "DB7093",
    "FFEFD5", "FFDAB9", "CD853F", "FFC0CB", "DDA0DD", "B0E0E6", "800080", "663399", "FF0000",
    "BC8F8F", "4169E1", "8B4513", "FA8072", "FAA460", "2E8B57", "FFF5EE", "A0522D", "C0C0C0",
    "87CEEB", "6A5ACD", "708090", "FFFAFA", "00FF7F", "4682B4", "D2B48C", "008080", "D8BFD8",
    "FF6347", "40E0D0", "EE82EE", "F5DEB3", "FFFFFF", "F5F5F5", "FFFF00", "9ACD32"
]

# Assign a unique color to each remote workspace
remote_workspace_colors = {}
color_index = 0

for row in df.itertuples():
    if pd.notna(row.RemoteWorkspace):
        remote_workspaces = row.RemoteWorkspace.split(',')
        for remote_workspace in remote_workspaces:
            remote_workspace = remote_workspace.strip()
            if remote_workspace not in remote_workspace_colors:
                remote_workspace_colors[remote_workspace] = colors[color_index % len(colors)]
                color_index += 1

# Apply the assigned colors to the corresponding cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.column_letter == 'D' and isinstance(cell.value, str):  # Check if cell value is a string
            remote_workspaces = cell.value.split(',')
            for remote_workspace in remote_workspaces:
                remote_workspace = remote_workspace.strip()
                for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for c in r:
                        if c.value == remote_workspace:
                            c.font = Font(color=remote_workspace_colors[remote_workspace])

# Also apply the colors to the 'RemoteWorkspace' column itself
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):  # Assuming 'RemoteWorkspace' is in column D
    for cell in row:
        if isinstance(cell.value, str):
            remote_workspaces = cell.value.split(',')
            colored_value = ""
            for remote_workspace in remote_workspaces:
                remote_workspace = remote_workspace.strip()
                color = remote_workspace_colors.get(remote_workspace, "000000")
                colored_value += f"{remote_workspace} "
                cell.font = Font(color=color)
            cell.value = colored_value.strip()

wb.save('workspaces_with_group_highlighted.xlsx')

print("Migration groups and highlights have been added to the Excel file.")